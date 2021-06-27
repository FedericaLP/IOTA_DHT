import argparse
import sys
from flask import Flask, request

from src.config import *
from src.node import Node, requests
from src.utils import INSERT, REMOVE, PIN_SEARCH, SUPERSET_SEARCH, reset_hops, get_hops

from flask_cors import CORS, cross_origin
from flask import jsonify
from openpyxl import load_workbook
from statistics import mean

app = Flask(APP_NAME)
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

################
RESULTS_FILE = 'results/test_results.xlsx'
SHEET_SUPERSET_PERF = 'superset_perf'
FIRST_COLUMN = 2

document = load_workbook(RESULTS_FILE)
superset_sheet_perf = document[SHEET_SUPERSET_PERF]


@app.route(INSERT)

def request_insert():
    print("argomenti", request)
    print("argomenti", request.args)
    #keyword = int(request.args.get('keyword'))
    keyword = request.args.get('keyword')
    obj = request.args.get('obj')
    print("KEY:", keyword, "OBJ:", obj)
    res = NODE.insert(keyword, obj)
    if type(res) is not str:
        res = res.text
  
    return res
    #return jsonify(res) #se non scommento quest dà errore lato javascript anche se il file viene inserito
    #return jsonify(({'success':True}), 200, {'ContentType':'application/json'} )



@app.route(REMOVE)
def request_remove():
    keyword = request.args.get('keyword')
    obj = request.args.get('obj')
    res = NODE.remove(keyword, obj)
    if type(res) is not str:
        res = res.text
    return res


@app.route(PIN_SEARCH)
def request_pin_search():
    keyword = request.args.get('keyword')
    threshold = request.args.get('threshold')
    print("key pin search", keyword)
    if threshold is None:
        res = NODE.pin_search(keyword)
    else:
        res = NODE.pin_search(keyword, int(threshold))
    if type(res) is not list:
        res = res.text
    else:
        res = ','.join(res)
    return res

superset_hops = [] ######
num_richieste = []#######
@app.route(SUPERSET_SEARCH)
def request_superset_search():

    
    keyword = request.args.get('keyword')
    threshold = int(request.args.get('threshold'))
    sender = request.args.get('sender')

    reset_hops()  

    res = NODE.superset_search(keyword, threshold, sender)
    superset_hops.append(get_hops() +1 ) #########
    num_richieste.append(1) ##########

    #print("hops IN superset:", get_hops() +1)###########

    print("stats", len(superset_hops), mean(superset_hops))
    superset_sheet_perf.cell(row=2, column=FIRST_COLUMN ).value = sum(num_richieste) #########
    superset_sheet_perf.cell(row=HYPERCUBE_SIZE, column=FIRST_COLUMN).value = mean(superset_hops) #########
    
    document.save(RESULTS_FILE)
    if type(res) is not list:
        res = res.text
    else:
        res = ','.join(res)
    return res


def parse_arguments(argv):
    parser = argparse.ArgumentParser()
    parser.add_argument('port', type=int, help='Port number to which connect the node')
    return parser.parse_args(argv)


if __name__ == '__main__':
    PORT = parse_arguments(sys.argv[1:]).port
    NODE_ID = PORT - INIT_PORT
    NODE = Node(NODE_ID)
    app.run(host=LOCAL_HOST, port=PORT, threaded=True)